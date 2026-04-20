"""
Ingestion script: parse every file under ./data and load into Qdrant.

Discovery is automatic — drop a file into ./data and re-run:
    python -m backend.ingest

Or from inside the backend container:
    docker compose exec backend python -m backend.ingest

Routing:
  - Filename (case-insensitive) contains a known keyword → specialized loader
    (preserves column semantics / aggregates large transaction tables).
  - Otherwise falls back to a generic loader by extension (.xlsx/.xls/.csv/.pdf).
  - Temp / lock / sqlite files are ignored.
"""
from __future__ import annotations

import csv
import sys
import time
import uuid
from pathlib import Path
from typing import Callable

import openpyxl
import pypdf
from google import genai
from google.genai.errors import ClientError
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, PointStruct, VectorParams

from .config import settings
from .vector_store import COLLECTION_NAME, EMBED_MODEL, VECTOR_DIM

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
# Tier 1 tested throughput: ~1500 docs/min (no sleep needed).
# Free tier users should set BATCH_SLEEP=9 to stay under 30k TPM.
BATCH_SIZE = 20            # items per embed call
BATCH_SLEEP = 0            # seconds between batches (set >0 only on free tier)
MAX_429_RETRIES = 5        # transient 429s still possible; exponential backoff
GENERIC_ROW_CAP = 500      # max rows ingested from an unrecognized spreadsheet
PDF_PAGE_CHAR_CAP = 1200   # truncate extracted page text to this many chars


# ---------------------------------------------------------------------------
# Embedding + upsert
# ---------------------------------------------------------------------------

def embed_batch(client: genai.Client, texts: list[str]) -> list[list[float]]:
    result = client.models.embed_content(model=EMBED_MODEL, contents=texts)
    return [e.values for e in result.embeddings]


def upsert_chunks(
    qdrant: QdrantClient,
    genai_client: genai.Client,
    chunks: list[dict],
) -> None:
    for i in range(0, len(chunks), BATCH_SIZE):
        batch = chunks[i : i + BATCH_SIZE]
        vectors = None
        for attempt in range(MAX_429_RETRIES + 1):
            try:
                vectors = embed_batch(genai_client, [c["content"] for c in batch])
                break
            except ClientError as e:
                # New google-genai surfaces the code as `e.code` (old versions used `e.status_code`).
                status = getattr(e, "code", None) or getattr(e, "status_code", None)
                if status == 429 and attempt < MAX_429_RETRIES:
                    retry_secs = 65 * (attempt + 1)  # 65s, 130s, 195s, 260s, 325s
                    print(f"  Rate limited ({status}) — sleeping {retry_secs}s (attempt {attempt + 1}/{MAX_429_RETRIES}) ...")
                    time.sleep(retry_secs)
                else:
                    raise
        if vectors is None:
            raise RuntimeError("embed_batch failed after all retries")
        points = [
            PointStruct(id=str(uuid.uuid4()), vector=vec, payload=chunk)
            for vec, chunk in zip(vectors, batch)
        ]
        qdrant.upsert(collection_name=COLLECTION_NAME, points=points)
        print(f"  {i + len(batch)}/{len(chunks)} chunks upserted")
        if i + BATCH_SIZE < len(chunks):
            time.sleep(BATCH_SLEEP)


# ---------------------------------------------------------------------------
# Specialized loaders — keep column-level semantics for known files
# ---------------------------------------------------------------------------

def load_item_specs(path: Path) -> list[dict]:
    """ItemSpecMaster — one chunk per SKU."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()

    chunks = []
    for row in rows[1:]:
        if not row[0]:
            continue
        item, desc, shelf_life, country, allergens, manufacturer, lead_time = (
            row[0], row[1] or "", row[10] or "", row[9] or "", row[14] or "", row[11] or "", row[12] or ""
        )
        chunks.append({
            "content": (
                f"Item: {item} | Description: {desc} | Shelf Life: {shelf_life} | "
                f"Country of Origin: {country} | Allergens: {allergens} | "
                f"Manufacturer: {manufacturer} | Lead Time: {lead_time}"
            ),
            "source": "ItemSpecMaster",
            "item_number": str(item),
        })
    print(f"  ItemSpecMaster        {len(chunks)} SKUs")
    return chunks


def load_vendor_master(path: Path) -> list[dict]:
    """VendorMaster — one chunk per vendor/brand."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()

    chunks = []
    for row in rows[1:]:
        if not row[0]:
            continue
        brand, product_line, category, vendor_id, status, country = (
            row[0] or "", row[1] or "", row[2] or "", row[4] or "", row[7] or "", row[8] or ""
        )
        skus, shipment_terms, payment_terms, currency = (
            row[11] or "", row[14] or "", row[15] or "", row[16] or ""
        )
        chunks.append({
            "content": (
                f"Brand: {brand} | Product Line: {product_line} | Category: {category} | "
                f"Vendor ID: {vendor_id} | Status: {status} | Country: {country} | "
                f"SKUs: {skus} | Shipment Terms: {shipment_terms} | "
                f"Payment Terms: {payment_terms} | Currency: {currency}"
            ),
            "source": "VendorMaster",
            "vendor_id": str(vendor_id),
            "brand": str(brand),
        })
    print(f"  VendorMaster          {len(chunks)} vendors")
    return chunks


def load_inventory(path: Path) -> list[dict]:
    """InventorySnapshot — combine SF + NJ per item."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    inventory: dict[str, dict] = {}
    for sheet_name in wb.sheetnames:
        site = "sf" if "1" in sheet_name or "SF" in sheet_name.upper() else "nj"
        for row in list(wb[sheet_name].iter_rows(values_only=True))[1:]:
            if not row[0]:
                continue
            item = str(row[0]).strip()
            desc = str(row[1]).strip() if row[1] else ""
            avail = int(row[2] or 0)
            on_hand = int(row[3] or 0)
            if item not in inventory:
                inventory[item] = {"desc": desc, "sf_avail": 0, "nj_avail": 0, "sf_on_hand": 0, "nj_on_hand": 0}
            inventory[item][f"{site}_avail"] += avail
            inventory[item][f"{site}_on_hand"] += on_hand
    wb.close()

    chunks = []
    for item, d in inventory.items():
        total = d["sf_avail"] + d["nj_avail"]
        chunks.append({
            "content": (
                f"Item: {item} | Description: {d['desc']} | "
                f"SF Available: {d['sf_avail']} | NJ Available: {d['nj_avail']} | "
                f"Total Available: {total} | SF On Hand: {d['sf_on_hand']} | NJ On Hand: {d['nj_on_hand']}"
            ),
            "source": "InventorySnapshot",
            "item_number": item,
        })
    print(f"  InventorySnapshot     {len(chunks)} items")
    return chunks


def load_sales_history(path: Path, top_n: int = 200) -> list[dict]:
    """SalesTransactionHistory.csv — aggregate by item, keep top N by quantity."""
    sales: dict[str, dict] = {}
    with open(path, encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            item = row.get("ITEMNMBR", "").strip()
            if not item:
                continue
            qty = float(row.get("QTYBSUOM") or 0)
            revenue = float(row.get("XTNDPRCE_adj") or 0)
            state = row.get("STATE", "").strip()
            if item not in sales:
                sales[item] = {"desc": row.get("ITEMDESC", "").strip(), "qty": 0.0, "revenue": 0.0, "states": set()}
            sales[item]["qty"] += qty
            sales[item]["revenue"] += revenue
            if state:
                sales[item]["states"].add(state)

    top = sorted(sales.items(), key=lambda x: x[1]["qty"], reverse=True)[:top_n]
    chunks = []
    for item, d in top:
        chunks.append({
            "content": (
                f"Item: {item} | Description: {d['desc']} | "
                f"Total Units Sold: {int(d['qty'])} | Total Revenue: ${d['revenue']:,.0f} | "
                f"Key States: {', '.join(sorted(d['states'])[:5])}"
            ),
            "source": "SalesHistory",
            "item_number": item,
        })
    print(f"  SalesHistory          {len(chunks)} top items (of {len(sales)} total)")
    return chunks


def load_purchase_orders(path: Path, top_vendors: int = 150) -> list[dict]:
    """PurchaseOrderHistory — aggregate by Vendor ID (top items shipped + spend)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    # Column layout: PO Number, PO Date, Required Date, Promised Ship Date, Receipt Date,
    # POP Receipt Number, Item Number, Item Description, QTY Shipped, QTY Invoiced,
    # Unit Cost, Extended Cost, Vendor ID, Location Code, Primary Ship To, Shipping Method
    by_vendor: dict[str, dict] = {}
    for row in rows[1:]:
        if not row or not row[12]:          # need Vendor ID
            continue
        vendor = str(row[12]).strip()
        item = str(row[6] or "").strip()
        desc = str(row[7] or "").strip()
        qty = float(row[8] or 0)
        spend = float(row[11] or 0)
        po_date = row[1]
        if vendor not in by_vendor:
            by_vendor[vendor] = {
                "pos": 0, "spend": 0.0, "items": {},
                "first": po_date, "last": po_date,
            }
        v = by_vendor[vendor]
        v["pos"] += 1
        v["spend"] += spend
        if item:
            agg = v["items"].setdefault(item, {"desc": desc, "qty": 0.0, "spend": 0.0})
            agg["qty"] += qty
            agg["spend"] += spend
        if po_date and (v["first"] is None or po_date < v["first"]):
            v["first"] = po_date
        if po_date and (v["last"] is None or po_date > v["last"]):
            v["last"] = po_date

    ranked = sorted(by_vendor.items(), key=lambda kv: kv[1]["spend"], reverse=True)[:top_vendors]
    chunks = []
    for vendor, v in ranked:
        top_items = sorted(v["items"].items(), key=lambda kv: kv[1]["spend"], reverse=True)[:5]
        items_str = "; ".join(
            f"{it} ({data['desc'][:40]}) qty={int(data['qty'])} ${data['spend']:,.0f}"
            for it, data in top_items
        )
        first = v["first"].strftime("%Y-%m") if v["first"] else "?"
        last = v["last"].strftime("%Y-%m") if v["last"] else "?"
        chunks.append({
            "content": (
                f"Vendor {vendor} | POs: {v['pos']} | Total Spend: ${v['spend']:,.0f} | "
                f"Active: {first} to {last} | Top items: {items_str}"
            ),
            "source": "PurchaseOrderHistory",
            "vendor_id": vendor,
        })
    print(f"  PurchaseOrderHistory  {len(chunks)} vendor summaries (of {len(by_vendor)} total)")
    return chunks


def load_transfers(path: Path, top_items: int = 200) -> list[dict]:
    """InternalTransferHistory — aggregate by Item Number (SF↔NJ movement)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # First sheet is the transaction list; extra sheets (TransferRate etc.) ignored.
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    # Column layout: Document Number, Document Date, Document Type, Item Number,
    # Item Description, U Of M, TRX QTY, Unit Cost, Extended Cost, TRX Location,
    # Transfer To Location, Document Status
    by_item: dict[str, dict] = {}
    for row in rows[1:]:
        if not row or not row[3]:
            continue
        item = str(row[3]).strip()
        desc = str(row[4] or "").strip()
        qty = float(row[6] or 0)
        cost = float(row[8] or 0)
        src = str(row[9] or "?").strip()
        dst = str(row[10] or "?").strip()
        route = f"{src}→{dst}"
        if item not in by_item:
            by_item[item] = {"desc": desc, "events": 0, "qty": 0.0, "cost": 0.0, "routes": {}}
        d = by_item[item]
        d["events"] += 1
        d["qty"] += qty
        d["cost"] += cost
        d["routes"][route] = d["routes"].get(route, 0) + 1

    ranked = sorted(by_item.items(), key=lambda kv: kv[1]["qty"], reverse=True)[:top_items]
    chunks = []
    for item, d in ranked:
        top_routes = sorted(d["routes"].items(), key=lambda kv: kv[1], reverse=True)[:3]
        routes_str = ", ".join(f"{r} ({n}x)" for r, n in top_routes)
        chunks.append({
            "content": (
                f"Item: {item} | Description: {d['desc']} | "
                f"Transfer events: {d['events']} | Total qty moved: {int(d['qty'])} | "
                f"Value: ${d['cost']:,.0f} | Routes: {routes_str}"
            ),
            "source": "InternalTransferHistory",
            "item_number": item,
        })
    print(f"  InternalTransferHistory {len(chunks)} items (of {len(by_item)} total)")
    return chunks


# ---------------------------------------------------------------------------
# Generic loaders — used for any file that doesn't match a specialized loader
# ---------------------------------------------------------------------------

def _is_readable(text: str) -> bool:
    if len(text) < 50:
        return False
    ascii_count = sum(1 for c in text if ord(c) < 128 and c.isprintable())
    return ascii_count / len(text) > 0.85


def load_generic_pdf(path: Path) -> list[dict]:
    """Any PDF — one chunk per readable page."""
    try:
        reader = pypdf.PdfReader(str(path))
    except Exception as exc:
        print(f"  {path.name}: pypdf failed ({exc!r})")
        return []
    source = path.stem
    chunks = []
    for i, page in enumerate(reader.pages):
        text = " ".join((page.extract_text() or "").split())
        if not _is_readable(text):
            continue
        chunks.append({
            "content": f"[{source} p{i + 1}] {text[:PDF_PAGE_CHAR_CAP]}",
            "source": source,
            "page": i + 1,
        })
    print(f"  {path.name}: {len(chunks)} readable pages")
    return chunks


def _format_row(headers: list[str], values: tuple) -> str:
    parts = []
    for h, v in zip(headers, values):
        if v is None or str(v).strip() == "":
            continue
        parts.append(f"{h}: {v}")
    return " | ".join(parts)


def load_generic_xlsx(path: Path, row_cap: int = GENERIC_ROW_CAP) -> list[dict]:
    """Any .xlsx/.xls — dump each row as `header: value | ...` (first sheet only)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[0])]
    source = path.stem
    chunks = []
    total = 0
    for values in rows[1:]:
        if not any(v not in (None, "") for v in values):
            continue
        total += 1
        if len(chunks) >= row_cap:
            continue
        content = _format_row(headers, values)
        if not content:
            continue
        chunks.append({
            "content": content,
            "source": source,
            "sheet": ws.title,
        })
    suffix = f" (capped at {row_cap}, full total {total})" if total > row_cap else ""
    print(f"  {path.name}: {len(chunks)} rows{suffix}")
    return chunks


def load_generic_csv(path: Path, row_cap: int = GENERIC_ROW_CAP) -> list[dict]:
    """Any .csv — dump each row as `header: value | ...`."""
    source = path.stem
    chunks = []
    total = 0
    with open(path, encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for row in reader:
            total += 1
            if len(chunks) >= row_cap:
                continue
            parts = [f"{h}: {row[h]}" for h in headers if row.get(h)]
            content = " | ".join(parts)
            if not content:
                continue
            chunks.append({"content": content, "source": source})
    suffix = f" (capped at {row_cap}, full total {total})" if total > row_cap else ""
    print(f"  {path.name}: {len(chunks)} rows{suffix}")
    return chunks


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

# (filename-substring, loader). First match wins. Substrings are lowercased.
SPECIALIZED: list[tuple[str, Callable[[Path], list[dict]]]] = [
    ("itemspec",          load_item_specs),
    ("vendormaster",      load_vendor_master),
    ("inventorysnapshot", load_inventory),
    ("salestransaction",  load_sales_history),
    ("purchaseorder",     load_purchase_orders),
    ("transfer",          load_transfers),
]

IGNORED_SUFFIXES = {".db", ".sqlite", ".sqlite3"}


def _is_ignored(path: Path) -> bool:
    name = path.name
    if name.startswith("~$") or name.startswith("."):
        return True
    if path.is_dir():
        return True
    if path.suffix.lower() in IGNORED_SUFFIXES:
        return True
    return False


def dispatch(path: Path) -> list[dict]:
    stem = path.stem.lower()
    for keyword, loader in SPECIALIZED:
        if keyword in stem:
            return loader(path)
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xls"}:
        return load_generic_xlsx(path)
    if ext == ".csv":
        return load_generic_csv(path)
    if ext == ".pdf":
        return load_generic_pdf(path)
    print(f"  {path.name}: skipped (no loader for extension {ext!r})")
    return []


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    api_key = settings.gemini_api_key
    qdrant_url = settings.qdrant_url or "http://localhost:6333"

    if not api_key:
        print("ERROR: GEMINI_API_KEY not set.")
        sys.exit(1)
    if not DATA_DIR.exists():
        print(f"ERROR: data dir not found at {DATA_DIR}")
        sys.exit(1)

    print(f"Connecting to Qdrant at {qdrant_url} ...")
    qdrant = QdrantClient(url=qdrant_url, timeout=10, check_compatibility=False)
    genai_client = genai.Client(api_key=api_key)

    existing = [c.name for c in qdrant.get_collections().collections]
    if COLLECTION_NAME in existing:
        qdrant.delete_collection(COLLECTION_NAME)
        print(f"Dropped existing '{COLLECTION_NAME}' collection")
    qdrant.create_collection(
        collection_name=COLLECTION_NAME,
        vectors_config=VectorParams(size=VECTOR_DIM, distance=Distance.COSINE),
    )
    print(f"Created '{COLLECTION_NAME}' collection\n")

    print(f"Scanning {DATA_DIR} ...")
    chunks: list[dict] = []
    for path in sorted(DATA_DIR.iterdir()):
        if _is_ignored(path):
            continue
        try:
            chunks += dispatch(path)
        except Exception as exc:
            print(f"  {path.name}: FAILED ({exc!r})")

    print(f"\nTotal: {len(chunks)} chunks — embedding and upserting ...")
    upsert_chunks(qdrant, genai_client, chunks)
    print("\nDone. PoP documents are now searchable.")


if __name__ == "__main__":
    main()
